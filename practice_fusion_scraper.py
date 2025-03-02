from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement

from pdf2image import convert_from_path
from PIL import Image
import pytesseract

import difflib
import datetime
import openpyxl
import PySimpleGUI as sg
import os
from time import sleep

from config import config
pytesseract.pytesseract.tesseract_cmd = config['tesseract_path']
COLUMN_NAME_MAPPING: dict = config['column_name_mapping']


# TODO: replace try-except-pass with suppress

def practice_fusion_login(driver: webdriver.Chrome):
    """Logs into practice fusion

    driver: selenium webpage driver
    """
    username_element: WebElement = driver.find_element(By.ID, config['pf_username_field_id'])
    username_element.send_keys(config['pf_username'])
    password_element: WebElement = driver.find_element(By.ID, config['pf_password_field_id'])
    password_element.send_keys(config['pf_password'])
    login_button_element: WebElement = driver.find_element(By.ID, config['pf_login_button_id'])
    login_button_element.click()

    # validate_login(driver)
    
    # allow new webpage to load
    sleep(1)

    # 2 factor authentication
    is_2fa_successful: bool = False
    try:
        send_code_button: WebElement = driver.find_element(By.ID, config['pf_send_2fa_code_button_id'])
        send_code_button.click()
        while not is_2fa_successful:
            code_entered: bool = enter_2fa_code(driver)
            if not code_entered:
                break

            is_2fa_successful = validate_2fa(driver)

    except:
        print('Invalid username/password')
        return False

    return is_2fa_successful


def enter_2fa_code(driver: webdriver.Chrome) -> bool:
    """Helps automate the 2fa process

    driver: selenium webpage driver
    return: if 2fa was entered
    """
    code: str = sg.popup_get_text(message='Enter two factor authentication code:', keep_on_top=True)
    
    if code is None or code == '':
        return False

    enter_code_element: WebElement = driver.find_element(By.ID, config['pf_enter_code_field_id'])
    enter_code_element.clear()
    enter_code_element.send_keys(code)
    send_code_button: WebElement = driver.find_element(By.ID, config['pf_send_code_button_id'])
    send_code_button.click()

    return True


def validate_2fa(driver: webdriver.Chrome) -> bool:
    """Checks if 2fa code was correct

    driver: selenium webpage driver
    return: if 2fa code was correct
    """
    try:
        sleep(.5)
        driver.find_element(By.XPATH, config['pf_wrong_code_xpath'])
    except:
        return True
    
    return False


def go_to_charts(driver: webdriver.Chrome):
    """Goes to the charts tab of practice fusion

    driver: selenium webpage driver
    """
    sidebar_elements: list[WebElement] = driver.find_elements(By.CLASS_NAME, config['pf_sidebar_label_class'])
    charts_label: WebElement = [element for element in sidebar_elements 
                                if element.text == config['pf_charts_element_text']][0]
    charts_label.click()


def get_date_of_birth(patient_data, row: int) -> str:
    """Gets the date of birth from the excel sheet

    patient_data: openpyxl excel sheet
    row: the row number of the current patient
    return: date of birth formatted as a string - mm/dd/yyyy
    """
    patient_dob: datetime.datetime = patient_data.cell(row, COLUMN_NAME_MAPPING['DOB']).value
    patient_dob_str: str = patient_dob.strftime('%m/%d/%Y')

    return patient_dob_str

def enter_date_of_birth(driver: webdriver.Chrome, dob: str):
    """Enters the date of birth in the practice fusion patient chart search bar

    driver: selenium webpage driver
    dob: date of birth as a string - mm/dd/yyyy
    """
    try:
        remove_dob_entry_element = driver.find_element(By.CSS_SELECTOR, config["pf_remove_dob_entry_css_select"])
    except:
        pass
    else:
        remove_dob_entry_element.click()
        sleep(.5)

    patient_search_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_patient_search_css_select'])
    patient_search_element.send_keys(f'{dob}')
    sleep(.5)
    patient_search_element.send_keys(Keys.RETURN)
    sleep(1)


def get_patient_full_name(patient_data, row: int) -> str:
    """Gets the first and last name of the patient from the excel file and concatenate them

    patient_data: openpyxl excel sheet
    row: the row number of the current patient
    return: full name of the patient - "<first_name> <last_name>"
    """
    patient_first_name: str = patient_data.cell(row, COLUMN_NAME_MAPPING['FIRST NAME']).value
    patient_last_name: str = patient_data.cell(row, COLUMN_NAME_MAPPING['LAST NAME']).value
    patient_full_name: str = f'{patient_first_name} {patient_last_name}'

    return patient_full_name


def navigate_to_patient_info(driver: webdriver.Chrome, full_name: str) -> bool:
    """Selects the patient with the closest matching name to the given full name

    driver: selenium webpage driver
    full_name: the full name of the patient - "<FirstName> <LastName>"
    return: if a matching patient was found
    """
    search_patient_first_names: list[WebElement] = driver.find_elements(
        By.CSS_SELECTOR, config['pf_first_name_css_select'])
    search_patient_last_names: list[WebElement] = driver.find_elements(
        By.CSS_SELECTOR, config['pf_last_name_css_select'])
    
    # gets list of patient names that match the DOB
    full_name_element_mapping: dict = {}
    for first_name_ele, last_name_ele in zip(search_patient_first_names, search_patient_last_names):
        search_first_name: str = first_name_ele.text.lower()
        search_last_name: str = last_name_ele.text.lower()
        search_full_name: str = f'{search_first_name} {search_last_name}'
        full_name_element_mapping[search_full_name] = first_name_ele

    # finds the one that is closest to the given name
    closest_matches: str = difflib.get_close_matches(
        word=full_name,
        possibilities=full_name_element_mapping.keys(),
        cutoff=0.75)
    if len(closest_matches) == 0:
        return False

    closest_match_element: WebElement = full_name_element_mapping[closest_matches[0]]
    closest_match_element.click()
    sleep(1) # wait for the webpage to load

    return True


def navigate_to_documents_tab(driver: webdriver.Chrome):
    """Navigates to the patient's document tab
    
    driver: selenium webpage driver
    """
    document_tab_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_documents_tab_css_select'])
    document_tab_element.click()
    sleep(1)

    # opens the signed vs unsigned document dropdown
    try:
        document_filter_dropdown: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_documents_filter_css_select'])
    except:
        driver.find_element(
            By.CSS_SELECTOR,
            'button[data-element="btn-rollback-patient"]',
        ).click()
        document_filter_dropdown: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_documents_filter_css_select'])

    document_filter_dropdown.click()
    sleep(.5)

    # selects the signed documents filter
    signed_documents_button: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_signed_documents_css_select'])
    signed_documents_button.click()
    sleep(1)


def navigate_to_profile_tab(driver: webdriver.Chrome):
    """Navigates the the patient's profile tab

    driver: selenium webpage driver
    """
    document_tab_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_profile_tab_css_select'])
    document_tab_element.click()
    sleep(1)


def get_patient_address(driver: webdriver.Chrome) -> tuple[str, str, str, str] | bool:
    """Gets the patient address from the patient profile

    driver: selenium webpage driver
    return: (address_line_1, address_line_2, city, state, zip_code) or False if no address is found
    """
    try:
        address_line_1: str = driver.find_element(By.CSS_SELECTOR, config['pf_address1_css_select']).text
        address_line_2: str = driver.find_element(By.CSS_SELECTOR, config['pf_address2_css_select']).text
        city_state_zip: str = driver.find_element(By.CSS_SELECTOR, config['pf_address3_css_select']).text
    except:
        return False

    city_state_zip = city_state_zip.replace(',', '')
    address_3_parts = city_state_zip.split(' ')
    state = address_3_parts[-2]
    zip_code = address_3_parts[-1]
    city = ' '.join(address_3_parts[0:-2])

    return address_line_1, address_line_2, city, state, zip_code


def get_patient_sex(driver: webdriver.Chrome):
    """Gets the patient's sex from the practice fusion header

    driver: selenium webpage driver
    """
    age_and_sex_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_sex_css_select'])
    sex: str = age_and_sex_element.text[-1]

    return sex


def navigate_to_desired_document(driver: webdriver.Chrome, operation: str, operation_date: datetime.datetime) -> bool:
    """Finds the document that matches the operation

    driver: selenium webpage driver
    operation: the operation type - used to find the corresponding document type
    operation_date: the date of the operation - used to filter results, searches for documents uploaded with 10 days of this date
    return: if a corresponding document was found
    """
    if operation is None:
        return False

    documents_container: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_document_container_css_select'])
    documents: list[WebElement] = documents_container.find_elements(By.XPATH, config['pf_document_row_xpath'])
    document_found: bool = False

    for document in documents:
        document_information_containers: list[WebElement] = document.find_elements(By.XPATH, config['xpath_child_selector'])
        
        # finds the document type (operation name)
        document_type: WebElement = document_information_containers[2]
        document_type_div: WebElement = document_type.find_element(By.XPATH, config['xpath_child_selector'])
        document_type_str: str = document_type_div.find_element(By.XPATH, config['xpath_child_selector']).text

        # skips if document type doesnt match
        if operation.strip().lower() not in document_type_str.lower():
            continue

        # finds the document upload date and converts to a string
        document_date_str: WebElement = document.find_elements(By.XPATH, config['xpath_child_selector'])[5].text
        document_date: datetime.datetime = datetime.datetime.strptime(document_date_str, '%m/%d/%Y')
        end_date_range: datetime.datetime = operation_date + datetime.timedelta(days=10) # date range to check

        # skips if document was not uploaded within the specified number of days
        if not (operation_date <= document_date <= end_date_range):
            continue

        # continues to the document page
        document_link_container: WebElement = document_information_containers[1]
        document_link_div: WebElement = document_link_container.find_element(By.XPATH, config['xpath_child_selector'])
        document_link_element: WebElement = document_link_div.find_element(By.XPATH, config['xpath_child_selector'])
        document_link_element.click()
        document_found = True
        break

    sleep(3)
    return document_found


def download_document(driver: webdriver.Chrome, patient_name: str, date: datetime.datetime, operation_type: str) -> str:
    """Downloads the operation document

    patient_name: the patient name - used to rename the downloaded files
    date: date of operation - used to rename the downloaded files
    operation_type: operation type - used to rename the downloaded files
    return: the name of the downloaded pdf file
    """
    if not os.path.exists(config['pf_pdf_download_location']):
        os.makedirs(config['pf_pdf_download_location'])

    print_buttons_container: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_print_buttons_css_select'])
    download_button: WebElement = print_buttons_container.find_element(By.XPATH, config['pf_download_button_xpath'])
    download_button.click()

    patient_info_tabs_container: WebElement = driver.find_element(By.CSS_SELECTOR, config['pf_tab_list_css_select'])
    pdf_name_element: WebElement = patient_info_tabs_container.find_element(By.XPATH, config['pf_pdf_name_xpath'])
    pdf_name: str = pdf_name_element.text

    operation_type = operation_type.strip().lower()
    date_str: str = date.strftime('%Y-%m-%d')
    new_file_name: str = f'{date_str}_{patient_name}_{operation_type}.pdf'

    old_file_path: str = f'{config["pf_pdf_download_location"]}{pdf_name}'
    new_file_path: str = f'{config["pf_pdf_download_location"]}{new_file_name}'

    sleep(1)

    if not os.path.exists(old_file_path):
        return ''

    if not os.path.exists(new_file_path):
        os.rename(old_file_path, new_file_path)

    return new_file_path


def get_text_from_pdf(pdf_file_path: str) -> str:
    """Converts the text in first page of the pdf to a string

    pdf_file_path: path to the pdf file
    return: the text in the first page of the pdf file
    """
    pdf_pages: list[Image.Image] = convert_from_path(pdf_file_path, 300)

    first_page: Image.Image = pdf_pages[0]
    first_page_text: str = pytesseract.image_to_string(first_page, lang='eng')

    return first_page_text


# TODO: support for multiple operations in one file
def parse_pdf_text(pdf_text: str) -> tuple[str, str, str]:
    """Gets the ID, procedure code, and diagnosis code from the pdf text

    pdf_text: the text in the pdf file
    return: (id, procedure_code, diagnosis_code)
    """
    pdf_text_by_line: list[str] = pdf_text.split('\n')

    id: str = ''
    procedure_code: str = ''
    diagnosis_code: str = ''
    for line in pdf_text_by_line:
        if line.startswith('ID:'):
            id_and_procedure: str = line.split(':')[1].strip()
            id = id_and_procedure.split(' ')[0]
        elif line.startswith('Procedure Code:'):
            proc_code_and_name: str = line.split(':')[1].strip()
            procedure_code = proc_code_and_name.split(' ')[0]
        elif line.startswith('Diagnosis code:'):
            diag_code_and_type: str = line.split(':')[1].strip()
            diagnosis_code = diag_code_and_type.split(' ')[0]

    return id, procedure_code, diagnosis_code


def close_patient_charts_tab(driver: webdriver.Chrome):
    """Closes the currently opened patient chart tab

    driver: selenium webpage driver
    """
    try:
        potential_patient_list_containers: list[WebElement] = driver.find_elements(By.CLASS_NAME, config['pf_close_charts_container_class'])
        patient_list_container: WebElement
        for container in potential_patient_list_containers:
            if container.text.startswith('Patient lists'):
                patient_list_container: WebElement = container
                break

        patient_list_container = patient_list_container.find_element(By.XPATH, config['xpath_child_selector'])
        patient_chart_container: WebElement = patient_list_container.find_elements(By.XPATH, config['xpath_child_selector'])[1]
        patient_chart_container: WebElement = patient_chart_container.find_element(By.XPATH, config['xpath_child_selector'])
        close_charts_button: WebElement = patient_chart_container.find_elements(By.XPATH, config['xpath_child_selector'])[2]
        close_charts_button.click()

        sleep(3)
    except:
        pass

    try:
        potential_close_dob_filter_button: list[WebElement] = driver.find_elements(By.CSS_SELECTOR, config['pf_dob_close_css_select'])
        close_dob_filter_button: WebElement
        for button in potential_close_dob_filter_button:
            if button.accessible_name == '\uf139':
                close_dob_filter_button = button
                break

        close_dob_filter_button.click()

        sleep(1)
    except:
        pass


# TODO: turn sleeps into sophisticated waits
# https://selenium-python.readthedocs.io/waits.html#explicit-waits
def get_all_patient_data():
    """Main driver function for getting all required patient data
    """
    # read the cleaned patient data
    workbook = openpyxl.load_workbook(f'data/{config["cleaned_workbook_name"]}')
    patient_data = workbook[config['main_worksheet_name']]
    
    # chrome options
    chrome_options = webdriver.ChromeOptions()
    prefs = {
        'download.default_directory': config['pf_pdf_download_location'], # setting download location
        'profile.default_content_settings.popups': False, # removes download prompt
    }
    chrome_options.add_experimental_option('prefs', prefs)
    chrome_options.add_argument('log-level=3') # removes warning/error logging from console
    driver = webdriver.Chrome(
        options=chrome_options,
    )
    driver.maximize_window()

    # loads the practice fusion webpage
    driver.get(config['pf_url'])

    login_successful: bool = practice_fusion_login(driver)
    if not login_successful:
        print('Login unsuccessful. Please check username, password, and two factor authentication code.')
        return

    sleep(3) # give time for webpage to load

    go_to_charts(driver)

    sleep(1) # give time for webpage to load

    for row_index in range(2, patient_data.max_row + 1):
        close_patient_charts_tab(driver)

        patient_dob: str = get_date_of_birth(patient_data, row_index)
        enter_date_of_birth(driver, patient_dob)
        
        patient_name: str = get_patient_full_name(patient_data, row_index)
        navigation_successful: bool = navigate_to_patient_info(driver, patient_name)
        if not navigation_successful:
            print(f'Unable to find patient: {patient_name} (excel row: {row_index})')
            continue

        navigate_to_profile_tab(driver)
        address_results = get_patient_address(driver)
        if address_results == False:
            print(f'Unable to find address for: {patient_name} (excel row: {row_index})')
            continue

        address_1, address_2, city, state, zip_code = address_results
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["ADDRESS LINE 1"], address_1)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["ADDRESS LINE 2"], address_2)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["CITY"], city)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["STATE"], state)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["ZIP CODE"], zip_code)

        sex: str = get_patient_sex(driver)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING['SEX'], sex)

        navigate_to_documents_tab(driver)
        operation_date: datetime.datetime = patient_data.cell(row_index, COLUMN_NAME_MAPPING['DATE']).value
        operation_type: str = patient_data.cell(row_index, COLUMN_NAME_MAPPING['PROCEDURE']).value
        navigation_successful = navigate_to_desired_document(driver, operation_type, operation_date)
        if not navigation_successful:
            print(f'Unable to find {operation_type} document for {patient_name} on {operation_date.strftime("%m/%d")} (excel row: {row_index})')
            continue

        pdf_file_path: str = download_document(driver, patient_name, operation_date, operation_type)
        if pdf_file_path == '':
            continue

        pdf_text: str = get_text_from_pdf(pdf_file_path)
        id, procedure_code, diagnosis_code = parse_pdf_text(pdf_text)

        patient_data.cell(row_index, COLUMN_NAME_MAPPING["PROCEDURE ID"], id)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["PROCEDURE CODE"], procedure_code)
        patient_data.cell(row_index, COLUMN_NAME_MAPPING["DIAGNOSIS"], diagnosis_code)

        workbook.save(f'data/{config["cleaned_workbook_name"]}')

    driver.close()
    print('Done with Practice Fusion scraping')


if __name__ == '__main__':
    get_all_patient_data()