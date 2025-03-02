from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from datetime import datetime
from time import sleep
import openpyxl
import PySimpleGUI as sg

from config import config
COLUMN_NAME_MAPPING: dict = config['column_name_mapping']

def system13_login(driver: webdriver.Chrome):
    """Logs into system13

    driver: selenium webpage driver
    """
    username_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_username_field_css_select'])
    username_element.send_keys(config['s13_username'])
    password_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_password_field_css_select'])
    password_element.send_keys(config['s13_password'])
    login_button_element: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_login_button_css_select'])
    login_button_element.click()
    
    # allow new webpage to load
    sleep(1)

    security_notice_button: WebElement = driver.find_element(By.CLASS_NAME, config['s13_security_notice_button_class'])
    security_notice_button.click()

    sleep(1)

    error_elements: list[WebElement] = driver.find_elements(By.CSS_SELECTOR, config['s13_login_error_css_select'])
    if len(error_elements) > 0:
        return False
    
    return True


def navigate_to_web_claim_entry(driver: webdriver.Chrome):
    """Navigates to the web claim entry page

    driver: selenium webpage driver
    """
    dashboard_button_container: WebElement = driver.find_element(By.ID, config['s13_dashboard_container_id'])
    dashboard_buttons: list[WebElement] = dashboard_button_container.find_elements(By.XPATH, config['xpath_child_selector'])
    web_claim_entry_button: WebElement = dashboard_buttons[0]
    web_claim_entry_button.click()

    sleep(1)


def system13_authentication(driver: webdriver.Chrome) -> bool:
    """Handles authentication for system13 login

    Args:
        driver: selenium webpage driver

    Returns:
        whether or not authentication was successfull
            false if the user did not enter an authentication code
    """
    enter_code_box: WebElement = driver.find_element(By.ID, config["s13_auth_code_box_id"])
    authentication_form_section: WebElement = driver.find_element(By.ID, config["s13_auth_form_section_id"])
    verify_button: WebElement = authentication_form_section.find_element(By.CSS_SELECTOR, config["s13_auth_send_button_css"])

    is_2fa_successful: bool = False
    while not is_2fa_successful:
        code = sg.popup_get_text("Enter two factor authentication code: ", keep_on_top=True)
        if code is None or code == "":
            break

        enter_code_box.send_keys(code)
        verify_button.click()
        sleep(.5)

        try:
            driver.find_element(By.CSS_SELECTOR, config["s13_auth_error_css"])
        except:
            is_2fa_successful = True

    return is_2fa_successful


def enter_in_search_box(driver: webdriver.Chrome, parent_element: WebElement,
                        value: str, extra_down_input: bool = False):
    """Enters text in a dropdown search box

    driver: selenium webpage driver
    parent_element: the inital dropdown element
    value: the string to enter in the dropdown
    extra_down_input: if an extra down input is needed to select the desired value (some strings have multiple results)
    """
    parent_element.send_keys(Keys.ENTER)
    
    search_bar: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_search_dropdown_css_select'])
    search_bar.send_keys(value)

    sleep(2)

    if extra_down_input:
        search_bar.send_keys(Keys.DOWN)
    search_bar.send_keys(Keys.ENTER)


def enter_patient_fields(driver: webdriver.Chrome, claim_information: dict):
    """Fills out the information needed for the patient information page

    driver: selenium webpage driver
    claim_information: openpyxl excel sheet
    """
    sleep(.5)
    # patient control number
    pcn_field: WebElement = driver.find_element(By.ID, config['s13_pcn_id'])
    pcn_field.send_keys(claim_information['id'])
    sleep(.5)

    # medical record number
    mrn_field: WebElement = driver.find_element(By.ID, config['s13_mrn_id'])
    mrn_field.send_keys(claim_information['id'])
    sleep(.5)

    # first name
    first_name_field: WebElement = driver.find_element(By.ID, config['s13_first_name_id'])
    first_name_field.send_keys(claim_information['first_name'])
    sleep(.5)

    # last name
    last_name_field: WebElement = driver.find_element(By.ID, config['s13_last_name_id'])
    last_name_field.send_keys(claim_information['last_name'])
    sleep(.5)

    # address line 1
    address_1_field: WebElement = driver.find_element(By.ID, config['s13_address_1_id'])
    address_1_field.send_keys(claim_information['address_1'])
    sleep(.5)

    # address line 2
    if claim_information['address_2'] is not None:
        address_2_field: WebElement = driver.find_element(By.ID, config['s13_address_2_id'])
        address_2_field.send_keys(claim_information['address_2'])
    sleep(.5)

    # city
    city_field: WebElement = driver.find_element(By.ID, config['s13_city_id'])
    city_field.send_keys(claim_information['city'])
    sleep(.5)

    # state
    state_field: WebElement = driver.find_element(By.ID, config['s13_state_id'])
    state_field.send_keys(claim_information['state'])
    sleep(.5)

    # zip code
    zip_code_field: WebElement = driver.find_element(By.ID, config['s13_zip_id'])
    zip_code_field.send_keys(claim_information['zip_code'])
    sleep(.5)

    # country
    # needs extra down input
    #   "united states" is a substring of "united states minor outlying islands"
    country_field: WebElement = driver.find_element(By.ID, config['s13_country_id'])
    country_field.send_keys(config['country'])
    sleep(.5)

    # ssn
    ssn_field: WebElement = driver.find_element(By.ID, config['s13_ssn_id'])
    ssn_field.send_keys(config['ssn'])
    sleep(.5)

    # sex
    # needs extra down input
    #   "male" is a substring of "female"
    sex_field: WebElement = driver.find_element(By.ID, config['s13_sex_id'])
    sex_field.send_keys(claim_information['sex'])
    sleep(.5)

    # ethnicity
    ethnicity_field: WebElement = driver.find_element(By.ID, config['s13_ethnicity_id'])
    ethnicity_field.send_keys(claim_information['ethnicity'])
    sleep(.5)

    # date of birth
    dob_field: WebElement = driver.find_element(By.ID, config['s13_dob_id'])
    dob_field.send_keys(claim_information['dob'])
    sleep(.5)

    # race
    race_field: WebElement = driver.find_element(By.ID, config['s13_race_id'])
    race_field.send_keys(claim_information['race'])
    sleep(.5)

    # date of operation
    bill_from_field: WebElement = driver.find_element(By.ID, config['s13_bill_from_id'])
    bill_from_field.send_keys(claim_information['date'])
    sleep(.5)

    # end date of operation (currently same as the date of operation)
    bill_to_field: WebElement = driver.find_element(By.ID, config['s13_bill_to_id'])
    sleep(.5)
    bill_to_field.send_keys(claim_information['date'])

    # claim frequency
    claim_frequency_field: WebElement = driver.find_element(By.ID, config['s13_frequency_id'])
    claim_frequency_field.send_keys(config['frequency_code'])
    sleep(.5)

    # admission type
    admission_type_field: WebElement = driver.find_element(By.ID, config['s13_admission_type_id'])
    admission_type_field.send_keys(config['admission_type'])
    sleep(.5)

    # patient origin
    origin_field: WebElement = driver.find_element(By.ID, config['s13_origin_id'])
    origin_field.send_keys(config['origin_code'])
    sleep(.5)

    # patient discharge status
    patient_status_field: WebElement = driver.find_element(By.ID, config['s13_patient_status_id'])
    patient_status_field.send_keys(config['patient_status_code'])
    sleep(.5)

    # facility type
    facility_field: WebElement = driver.find_element(By.ID, config['s13_facility_id'])
    facility_field.send_keys(config['facility_code'])


def enter_payers_fields(driver: webdriver.Chrome, claim_entry_information: dict):
    """Fills out the information for the payers page

    driver: selenium webpage driver
    claim_information: openpyxl excel sheet
    """
    # primary payer type
    primary_payer_type_field: WebElement = driver.find_element(By.ID, config['s13_primary_payer_type_id'])
    primary_payer_type_field.click()

    primary_payer_type_dropdown: WebElement = driver.find_element(By.ID, config["s13_primary_payer_dropdown_id"])
    primary_payer_type_dropdown_options: list[WebElement] = primary_payer_type_dropdown.find_elements(By.CSS_SELECTOR, "li")

    insurance_type: float | str = claim_entry_information['insurance_type']
    if isinstance(insurance_type, float):
        insurance_type = str(int(insurance_type))

    dropdown_option_to_click: WebElement | None = None
    for dropdown_option in primary_payer_type_dropdown_options:
        if dropdown_option.text.startswith(insurance_type):
            dropdown_option_to_click = dropdown_option
            break

    if dropdown_option_to_click is None:
        print(f"No matching option for Primary Payer Type for {claim_entry_information['first_name']} {claim_entry_information['last_name']}")
        return

    dropdown_option.click()

    # primary payer name
    primary_payer_name_field: WebElement = driver.find_element(By.ID, config['s13_primary_payer_name_id'])
    primary_payer_name_field.send_keys(claim_entry_information['insurance_name'])


def enter_charges_fields(driver: webdriver.Chrome, claim_entry_information: dict):
    """Fills out the information needed for the charges fields

    driver: selenium webpage driver
    claim_information: openpyxl excel sheet
    """
    # TODO: find out why the charges tab dies when using the normal navigate to next section method
    payers_tab = driver.find_element(By.ID, config["s13_payers_tab_id"])
    payers_tab.click()
    sleep(1)

    charges_tab = driver.find_element(By.ID, config["s13_charges_tab_id"])
    charges_tab.click()
    sleep(1)

    # revenue code
    revenue_code_field: WebElement = driver.find_element(By.ID, config['s13_revenue_code_id'])
    revenue_code_field.send_keys(config['s13_revenue_code'])
    sleep(.5)

    # procedure code
    proceedure_code_field: WebElement = driver.find_element(By.ID, config['s13_procedure_code_id'])
    enter_in_search_box(driver, proceedure_code_field, claim_entry_information['procedure_code'], False)
    sleep(.5)

    # procedure date
    procedure_date_field: WebElement = driver.find_element(By.ID, config['s13_procedure_date_id'])
    procedure_date_field.send_keys(claim_entry_information['date'])
    sleep(.5)

    # end date of procedure (currently same as the date of operation)
    procedure_date_to_field: WebElement = driver.find_element(By.ID, config['s13_procedure_to_date_id'])
    sleep(.5)
    procedure_date_to_field.send_keys(claim_entry_information['date'])

    # unit rate
    rate_field: WebElement = driver.find_element(By.ID, config['s13_unit_rate_id'])
    if claim_entry_information['procedure_code'] == config['egd_code']:
        rate_field.send_keys(config['s13_edg_charge'])
    else:
        rate_field.send_keys(config['s13_colon_charge'])
    sleep(.5)

    # quantity
    quantity_field: WebElement = driver.find_element(By.ID, config['s13_quantity_id'])
    quantity_field.send_keys(config['s13_charge_quantity'])
    sleep(1)

    # unit
    # "un" is a substring of "international unit"
    units_field: WebElement = driver.find_element(By.ID, config['s13_unit_id'])
    units_field.send_keys(config['s13_charge_unit'])
    sleep(.5)

    # charge qualifier
    qualifier_field: WebElement = driver.find_element(By.ID, config['s13_charge_qualifier_id'])
    qualifier_field.send_keys(config['s13_qualifier_code'])


def enter_diagnoses_fields(driver: webdriver.Chrome, diagnosis_code: str):
    """Fills out the information needed for the diagnoses page

    Only does the primary diagnosis

    driver: selenium webpage driver
    claim_information: openpyxl excel sheet
    """
    sleep(.5)

    diagnosis_code = diagnosis_code.replace('.', '')

    principal_diagnosis_field: WebElement = driver.find_element(By.ID, config['s13_principal_diagnosis_id'])
    enter_in_search_box(driver, principal_diagnosis_field, diagnosis_code, False)
    sleep(.5)

    visit_reason_field: WebElement = driver.find_element(By.ID, config['s13_visit_reason_id'])
    enter_in_search_box(driver, visit_reason_field, diagnosis_code, False)


def enter_practitioners_fields(driver: webdriver.Chrome):
    """Fills out the information needed for the practitioners page

    driver: selenium webpage driver
    claim_information: openpyxl excel sheet
    """
    # practitioner id
    practitioner_id_field: WebElement = driver.find_element(By.ID, config['s13_practitioner_id_id'])
    practitioner_id_field.send_keys(config['practitioner_id'])
    sleep(.5)

    # practitioner first name
    practitioner_first_name_field: WebElement = driver.find_element(By.ID, config['s13_practitioner_first_name_id'])
    practitioner_first_name_field.send_keys(config['practitioner_first_name'])
    sleep(.5)

    # practitioner last name
    practitioner_last_name_field: WebElement = driver.find_element(By.ID, config['s13_practitioner_last_name_id'])
    practitioner_last_name_field.send_keys(config['practitioner_last_name'])
    sleep(.5)

    # practitioner type
    practitioner_type_field: WebElement = driver.find_element(By.ID, config['s13_practitioner_type_id'])
    practitioner_type_field.send_keys(config['practitioner_type'])


def navigate_to_next_section(driver: webdriver.Chrome):
    """Navigates to the next page/section

    driver: selenium webpage driver
    """
    sleep(1)
    
    next_section_button: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_next_section_css_select'])
    next_section_button.click()

    sleep(1)


def submit_claim(driver: webdriver.Chrome):
    """Submits the claim for error checking

    driver: selenium webpage driver
    """
    sleep(1)
    
    check_for_errors_button: WebElement = driver.find_element(By.CSS_SELECTOR, config['s13_check_for_errors_css_select'])
    check_for_errors_button.click()

    sleep(1)


def add_new_claim(driver: webdriver.Chrome):
    """Starts a new claim

    driver: selenium webpage driver
    """
    sleep(1)

    navigation_buttons_container: WebElement = driver.find_element(By.ID, config['s13_new_claim_button_container_id'])
    add_new_claim_button: WebElement = navigation_buttons_container.find_element(By.XPATH, config['s13_new_claim_xpath'])
    add_new_claim_button.click()

    sleep(1)


# TODO: turn sleeps into sophisticated waits
# https://selenium-python.readthedocs.io/waits.html#explicit-waits
def enter_all_patient_data():
    """Main driver function for entering all patient claims data
    """
    workbook = openpyxl.load_workbook(f'data/{config["cleaned_workbook_name"]}')
    patient_data = workbook[config['main_worksheet_name']]

    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('log-level=3')
    driver = webdriver.Chrome(
        options=chrome_options,
    )
    driver.maximize_window()
    driver.get(config['s13_url'])
    successful_login: bool = system13_login(driver)
    if not successful_login:
        print('Invalid Username or Password. Please check login credentials and try again.')
        return
    navigate_to_web_claim_entry(driver)
    system13_authentication(driver)

    for row_index in range(2, patient_data.max_row + 1):
        date: datetime = patient_data.cell(row_index, COLUMN_NAME_MAPPING['DATE']).value
        dob: datetime = patient_data.cell(row_index, COLUMN_NAME_MAPPING['DOB']).value
        claim_entry_information: dict = {
            'date': date.strftime('%m%d%Y'),
            'first_name': patient_data.cell(row_index, COLUMN_NAME_MAPPING['FIRST NAME']).value,
            'last_name': patient_data.cell(row_index, COLUMN_NAME_MAPPING['LAST NAME']).value,
            'id': patient_data.cell(row_index, COLUMN_NAME_MAPPING['PROCEDURE ID']).value,
            'procedure_code': patient_data.cell(row_index, COLUMN_NAME_MAPPING['PROCEDURE CODE']).value,
            'dob': dob.strftime('%m%d%Y'),
            'insurance_type': patient_data.cell(row_index, COLUMN_NAME_MAPPING['INSURANCE TYPE']).value,
            'insurance_name': patient_data.cell(row_index, COLUMN_NAME_MAPPING['INSURANCE']).value,
            'address_1': patient_data.cell(row_index, COLUMN_NAME_MAPPING['ADDRESS LINE 1']).value,
            'address_2': patient_data.cell(row_index, COLUMN_NAME_MAPPING['ADDRESS LINE 2']).value,
            'city': patient_data.cell(row_index, COLUMN_NAME_MAPPING['CITY']).value,
            'state': patient_data.cell(row_index, COLUMN_NAME_MAPPING['STATE']).value,
            'zip_code': patient_data.cell(row_index, COLUMN_NAME_MAPPING['ZIP CODE']).value,
            'sex': patient_data.cell(row_index, COLUMN_NAME_MAPPING['SEX']).value,
            'ethnicity': patient_data.cell(row_index, COLUMN_NAME_MAPPING['ETHNICITY']).value,
            'race': patient_data.cell(row_index, COLUMN_NAME_MAPPING['RACE']).value,
            'diagnosis_code': patient_data.cell(row_index, COLUMN_NAME_MAPPING['DIAGNOSIS']).value
        }

        empty_fields: list[str] = [key for key, value in claim_entry_information.items() if value is None]
        if empty_fields != ['address_2'] and empty_fields != []:
            continue

        enter_patient_fields(driver, claim_entry_information)
        navigate_to_next_section(driver)

        enter_payers_fields(driver, claim_entry_information)
        navigate_to_next_section(driver)

        enter_charges_fields(driver, claim_entry_information)
        navigate_to_next_section(driver)

        enter_diagnoses_fields(driver, claim_entry_information['diagnosis_code'])
        navigate_to_next_section(driver)

        enter_practitioners_fields(driver)
        submit_claim(driver)

        add_new_claim(driver)

    driver.close()
    print('Done with System13 entry')

if __name__ == '__main__':
    enter_all_patient_data()
