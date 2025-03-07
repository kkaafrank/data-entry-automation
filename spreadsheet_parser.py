import re
import openpyxl
import datetime
import PySimpleGUI as psg

from config import config

def remove_unnecessary_rows(worksheet):
    """Removes empty and unwanted rows from the excel sheet

    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    confirmed_keywords: list[str] = config['confirmed_appointment_keywords'].split(',')

    for row_index in range(worksheet.max_row, 1, -1):
        date_cell = worksheet.cell(row_index, config['column_name_mapping']['DATE'])
        patient_name_cell = worksheet.cell(row_index, config['column_name_mapping']['FIRST NAME'])

        row_doesnt_have_date: bool = date_cell.value is None
        row_doesnt_have_patient: bool = patient_name_cell.value is None
        if row_doesnt_have_date or row_doesnt_have_patient:
            worksheet.delete_rows(row_index, 1)
            continue

        # confirmed: bool = any([confirm_keyword in patient_name_cell.value.lower() for confirm_keyword in confirmed_keywords])
        # if not confirmed:
        #     worksheet.delete_rows(row_index, 1)
        #     continue

        is_monday: bool = date_cell.value.weekday() == 0
        is_wednesday: bool = date_cell.value.weekday() == 2
        if not is_monday and not is_wednesday:
            worksheet.delete_rows(row_index, 1)

    return worksheet


def edit_spreadsheet_columns(worksheet):
    """Removes unnecessary columns and adds empty ones that will be used later
    
    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    # remove unnecessary columns
    for col_index in range(worksheet.max_column, 0, -1):
        if worksheet.cell(1, col_index).value == 'PT NAME':
            worksheet.cell(1, col_index, 'FIRST NAME')
            continue

        if worksheet.cell(1, col_index).value not in config['column_name_mapping'].keys():
            worksheet.delete_cols(col_index, 1)
    print('    Done removing columns')

    sheet_column_names: list[str] = [worksheet.cell(1, col_index).value for col_index in range(1, len(config['column_name_mapping']) + 1)]
    sheet_column_names = [col_name for col_name in sheet_column_names if col_name is not None]
    # add empty columns that will be used later
    for col_name, col_index in config['column_name_mapping'].items():
        if col_name not in sheet_column_names:
            worksheet.insert_cols(col_index)
            worksheet.cell(1, col_index, col_name)
            # sheet_column_names.append(col_name)
    print('    Done adding columns')
        
    return worksheet


def clean_patient_name_cells(worksheet):
    """Attempts to remove all extra information in the patient name cell

    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    for row_index in range(worksheet.max_row, 1, -1):
        patient_name_cell_value: str = worksheet.cell(row_index, config['column_name_mapping']['FIRST NAME']).value
        if patient_name_cell_value is None:
            continue

        patient_name_cell_value = patient_name_cell_value.lower()
        patient_name_cell_value = patient_name_cell_value.replace('comfirmed', 'confirmed')

        patient_cell_parts = re.split(' |\(|,', patient_name_cell_value, maxsplit=3)
        for substring_index in range(len(patient_cell_parts) -1, -1, -1):
            substring = patient_cell_parts[substring_index]

            if 'confirmed' in substring:
                patient_cell_parts.remove(substring)
                continue
            if substring.startswith('-'):
                patient_cell_parts.remove(substring)
                continue
            if substring.endswith('-'):
                patient_cell_parts[substring_index] = substring[:-1]
            if not re.match(r'[A-Za-z\-]+', substring):
                patient_cell_parts.remove(substring)
                continue
        
        first_name: str = ' '.join(patient_cell_parts[:-1])
        last_name: str = patient_cell_parts[-1]
        worksheet.cell(row_index, config['column_name_mapping']['FIRST NAME'], first_name)
        worksheet.cell(row_index, config['column_name_mapping']['LAST NAME'], last_name)

    return worksheet


def separate_combined_procedures(worksheet):
    """Separates a patient with two procedures (colon and egd)
    
    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    for row_index in range(worksheet.max_row, 1, -1):
        procedure: str = worksheet.cell(row_index, config['column_name_mapping']['PROCEDURE']).value
        if procedure is None:
            continue
        procedure = procedure.lower()
        is_colonoscopy: bool = 'colon' in procedure
        is_egd: bool = 'egd' in procedure

        if not is_colonoscopy or not is_egd:
            continue

        worksheet.insert_rows(row_index + 1, 1)
        procedure = procedure.replace('egd', '').replace('/', '')
        worksheet.cell(row_index, config['column_name_mapping']['PROCEDURE'], procedure)
        for col_index in range(1, worksheet.max_column + 1, 1):
            if col_index == config['column_name_mapping']['PROCEDURE']:
                worksheet.cell(row_index + 1, col_index, 'egd')
            elif col_index == config['column_name_mapping']['DATE']:
                date: datetime.datetime = worksheet.cell(row_index, col_index).value
                worksheet.cell(row_index + 1, col_index, date)
                worksheet.cell(row_index + 1, col_index).number_format = 'mm-dd-yy'
            else:
                original_row_value = worksheet.cell(row_index, col_index).value
                worksheet.cell(row_index + 1, col_index, original_row_value)

    return worksheet


def derive_insurance_type(worksheet):
    """Derives insurance types from the insurance string

    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    self_pay_spellings: list[str] = config['self_pay_spellings'].split(',')
    for row_index in range(worksheet.max_row, 1, -1):
        insurace_provider: str = worksheet.cell(row_index, config['column_name_mapping']['INSURANCE']).value
        if insurace_provider is None:
            insurace_provider = 'self-pay'
        else:
            insurace_provider = insurace_provider.lower()
        insurance_type: str = ''

        if 'bcbs' in insurace_provider:
            insurance_type = 'BL'
        elif any([self_pay_spelling in insurace_provider for self_pay_spelling in self_pay_spellings]):
            insurance_type = 'Z'
        elif 'ppo' in insurace_provider:
            insurance_type = '12'
        else:
            insurance_type = '16'
        
        worksheet.cell(row_index, config['column_name_mapping']['INSURANCE TYPE'], insurance_type)

    return worksheet


def strip_name_fields(worksheet):
    """Removes leading and trailling spaces in the name fields

    worksheet: openpyxl worksheet
    return: edited openpyxl worksheet
    """
    for row_index in range(worksheet.max_row, 1, -1):
        first_name: str = worksheet.cell(row_index, config['column_name_mapping']['FIRST NAME']).value
        last_name: str = worksheet.cell(row_index, config['column_name_mapping']['LAST NAME']).value

        first_name = first_name.strip()
        last_name = last_name.strip()

        worksheet.cell(row_index, config['column_name_mapping']['FIRST NAME'], first_name)
        worksheet.cell(row_index, config['column_name_mapping']['LAST NAME'], last_name)

    return worksheet


def create_file_selection_layout() -> list[list[psg.Element]]:
    file_prompt_text = psg.Text(
        text='Select Initial Spreadsheet:',
    )
    file_input = psg.Input(
        key=config['spreadsheet_input_key'],
        enable_events=True,
    )
    file_browse_button = psg.FileBrowse(
        button_text='Browse Files',
        file_types=(('Excel Spreadsheets', '.xlsx'), ('All Files', '.*')),
    )
    file_gui_layout = [
        [file_prompt_text],
        [file_input, file_browse_button],
        [psg.Submit(key=config['ok_key']), psg.Cancel(config['cancel_key'])]
    ]
    return file_gui_layout

def get_initial_spreadsheet_file() -> str:
    layout = create_file_selection_layout()
    window = psg.Window(
        title='Select Spreadsheet File',
        layout=layout,
    )

    ok_key: str = config['ok_key']
    cancel_key: str = config['cancel_key']

    while True:
        event, values = window.read()

        if event == psg.WIN_CLOSED or event == cancel_key:
            window.Close()
            return None
        
        if event == ok_key:
            window.Close()
            break

    file_name: str = values[config['spreadsheet_input_key']]
    return file_name


def check_valid_excel_name(file_name: str) -> bool:
    if not file_name:
        return False
    elif '.xlsx' not in file_name:
        return False
    
    return True


def parse_spreadsheet():
    """Main driver function for the spreadsheet cleanup and parsing routine
    """
    file_name: str = get_initial_spreadsheet_file()
    is_valid_file_name: bool = check_valid_excel_name(file_name)
    if not is_valid_file_name:
        return

    workbook = openpyxl.load_workbook(file_name)
    # remove unneeded sheets
    for sheet_name in workbook.sheetnames:
        if sheet_name != config['main_worksheet_name']:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
    worksheet = workbook[config['main_worksheet_name']]

    # data cleaning
    worksheet = edit_spreadsheet_columns(worksheet)
    print('Done editing columns')
    worksheet = remove_unnecessary_rows(worksheet)
    print('Done removing rows')
    worksheet = clean_patient_name_cells(worksheet)
    print('Done cleaning patient names')
    worksheet = separate_combined_procedures(worksheet)
    print('Done separating colon and egd procedures')
    worksheet = derive_insurance_type(worksheet)
    print('Done deriving insurance types')
    worksheet = strip_name_fields(worksheet)
    print('Done with excel cleanup')

    worksheet.page_setup.fitToWidth = 1
    workbook.save(f'data/{config["cleaned_workbook_name"]}')


if __name__ == '__main__':
    parse_spreadsheet()
